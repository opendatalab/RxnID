import os
import re
import json
import csv
import argparse
import logging
from datetime import datetime
from collections import defaultdict, Counter
from typing import List, Dict, Tuple, Optional

import cv2
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from scipy.optimize import linear_sum_assignment
import matplotlib.pyplot as plt
import matplotlib.patches as patches


# =========================
# 全局阈值
# =========================
COND_TEXT_EDIT_DIST_THRES = 0.20  # 条件文本允许的最大归一化编辑距离


# =========================
# 日志
# =========================
def setup_logging(output_dir: str) -> str:
    """在 output_dir/logs 下创建日志文件，并同时输出到控制台。"""
    log_dir = os.path.join(output_dir, "logs")
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(log_dir, f"evaluation_{timestamp}.log")
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler()],
    )
    return log_file


# =========================
# 文本编辑距离（纯 Python）
# =========================
def levenshtein_distance(a: str, b: str) -> int:
    """标准 Levenshtein 距离的简洁实现。适用于 <= 2k 字符的场景。"""
    if a == b:
        return 0
    la, lb = len(a), len(b)
    if la == 0:
        return lb
    if lb == 0:
        return la

    prev = list(range(lb + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        for j, cb in enumerate(b, 1):
            curr.append(prev[j - 1] if ca == cb else 1 + min(prev[j - 1], prev[j], curr[-1]))
        prev = curr
    return prev[-1]


def normalized_edit_distance(a: str, b: str) -> float:
    """归一化编辑距离 ∈ [0,1]；0=完全相同。"""
    return levenshtein_distance(a, b) / max(len(a), len(b), 1)


# =========================
# 调试与数据检查
# =========================
def debug_filename_matching(gold_images: List[Dict], pred_images: List[Dict], max_examples: int = 5) -> None:
    """展示文件名样例，并用编辑距离检查高度相似但不相等的文件名。"""
    print("\n=== Filename Matching Debug ===")
    gold_filenames = [img.get("file_name", "") for img in gold_images]
    pred_filenames = [img.get("file_name", "") for img in pred_images]
    gold_set, pred_set = set(gold_filenames), set(pred_filenames)

    print(f"Sample gold filenames ({min(max_examples, len(gold_set))}):")
    for i, fn in enumerate(sorted(gold_set)[:max_examples], 1):
        print(f"  {i}. {fn}")

    print(f"\nSample pred filenames ({min(max_examples, len(pred_set))}):")
    for i, fn in enumerate(sorted(pred_set)[:max_examples], 1):
        print(f"  {i}. {fn}")

    print("\nChecking for similar filenames (>=80% similarity)...")
    similar_pairs = []
    for g in gold_set:
        for p in pred_set:
            if g != p:
                sim = 1 - normalized_edit_distance(g, p)
                if sim >= 0.8:
                    similar_pairs.append((g, p, sim))

    if similar_pairs:
        similar_pairs.sort(key=lambda x: x[2], reverse=True)
        for g, p, s in similar_pairs[:10]:
            print(f"  {g} <-> {p} (similarity: {s:.3f})")
    else:
        print("No similar filenames found.")
    print("=== End Debug ===\n")


def validate_data_integrity(gold_images: List[Dict], pred_images: List[Dict]) -> None:
    """基本结构检查、缺失字段与计数统计。"""
    print("\n=== Data Integrity Validation ===")

    def find_duplicates(names: List[str]) -> List[str]:
        cnt = Counter(names)
        return [n for n, c in cnt.items() if c > 1]

    gold_names = [img.get("file_name", "") for img in gold_images]
    pred_names = [img.get("file_name", "") for img in pred_images]

    gold_duplicates = find_duplicates(gold_names)
    pred_duplicates = find_duplicates(pred_names)
    if gold_duplicates:
        print(f"Warning: Duplicate filenames in gold: {len(gold_duplicates)} (e.g., {gold_duplicates[:3]})")
    if pred_duplicates:
        print(f"Warning: Duplicate filenames in pred: {len(pred_duplicates)} (e.g., {pred_duplicates[:3]})")

    def missing_fields(items: List[Dict], tag: str) -> None:
        misses = []
        for i, img in enumerate(items):
            if "file_name" not in img:
                misses.append(f"{tag}[{i}] missing file_name")
            if "bboxes" not in img:
                misses.append(f"{tag}[{i}] missing bboxes")
            if "reactions" not in img:
                misses.append(f"{tag}[{i}] missing reactions")
        if misses:
            print(f"Warning: {len(misses)} missing fields in {tag} (showing up to 5):")
            for m in misses[:5]:
                print(" ", m)

    missing_fields(gold_images, "gold")
    missing_fields(pred_images, "pred")

    total_gold_reacts = sum(len(img.get("reactions", [])) for img in gold_images)
    total_pred_reacts = sum(len(img.get("reactions", [])) for img in pred_images)
    total_gold_boxes = sum(len(img.get("bboxes", [])) for img in gold_images)
    total_pred_boxes = sum(len(img.get("bboxes", [])) for img in pred_images)

    print("\nData Statistics:")
    print(f"  Gold images: {len(gold_images)}, reactions: {total_gold_reacts}, bboxes: {total_gold_boxes}")
    print(f"  Pred images: {len(pred_images)}, reactions: {total_pred_reacts}, bboxes: {total_pred_boxes}")
    print("=== End Validation ===\n")


# =========================
# 评估核心
# =========================
class BBox:
    """标准化 bbox 为 [0,1] 归一化坐标，并携带分类与文本（若有）。"""

    def __init__(self, bbox_data: Dict, image_data: "ReactionImageData", xyxy: bool = False, normalized: bool = False):
        self.data = bbox_data
        self.image_data = image_data
        self.category_id = bbox_data.get("category_id", 1)  # 1=mol；默认当作分子

        self.width = getattr(image_data, "width", 1) or 1
        self.height = getattr(image_data, "height", 1) or 1

        if xyxy:
            x1, y1, x2, y2 = bbox_data["bbox"]
        else:
            x1, y1, w, h = bbox_data["bbox"]
            x2, y2 = x1 + w, y1 + h

        if not normalized:
            x1, y1, x2, y2 = x1 / self.width, y1 / self.height, x2 / self.width, y2 / self.height

        self.x1, self.y1, self.x2, self.y2 = x1, y1, x2, y2
        self.text = (bbox_data.get("text") or "").strip()

    @property
    def is_mol(self) -> bool:
        return self.category_id == 1

    def to_pixel(self) -> List[float]:
        x1 = self.x1 * self.width
        y1 = self.y1 * self.height
        x2 = self.x2 * self.width
        y2 = self.y2 * self.height
        return [x1, y1, x2 - x1, y2 - y1]


class Reaction:
    """单条反应的三元组（reactants/conditions/products）及比较逻辑。"""

    def __init__(self, reaction_data: Dict, bboxes_source: List[BBox], image_data: "ReactionImageData", filter_recover: bool = False):
        self.image_data = image_data
        self.filter_recover = filter_recover
        self.bboxes: List[BBox] = []
        self.reactants = self._process_role(reaction_data.get("reactants", []), bboxes_source)
        self.conditions = self._process_role(reaction_data.get("conditions", []), bboxes_source)
        self.products = self._process_role(reaction_data.get("products", []), bboxes_source)

    @staticmethod
    def _normalize_text(txt: str) -> str:
        """面向化学文本的轻量归一化：全/半角、上下标、单位/符号、空白删减、转小写等。"""
        if not txt:
            return ""
        txt = txt.strip()

        # 全角 → 半角
        txt = txt.translate(str.maketrans("０１２３４５６７８９", "0123456789"))
        txt = txt.translate(str.maketrans("ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ", "ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
        txt = txt.translate(str.maketrans("ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ", "abcdefghijklmnopqrstuvwxyz"))
        txt = txt.translate(str.maketrans("＋－×÷＝（）【】｛｝［］", "+-*/=()[]{}[]"))

        # 下标 / 上标字符
        sub_map = {"₀": "0", "₁": "1", "₂": "2", "₃": "3", "₄": "4", "₅": "5", "₆": "6", "₇": "7", "₈": "8", "₉": "9",
                   "ₐ": "a", "ₑ": "e", "ᵢ": "i", "ₒ": "o", "ᵣ": "r", "ᵤ": "u", "ᵥ": "v", "ₓ": "x"}
        sup_map = {"⁰": "0", "¹": "1", "²": "2", "³": "3", "⁴": "4", "⁵": "5", "⁶": "6", "⁷": "7", "⁸": "8", "⁹": "9",
                   "ᵃ": "a", "ᵇ": "b", "ᶜ": "c", "ᵈ": "d", "ᵉ": "e", "ᶠ": "f", "ᵍ": "g", "ʰ": "h", "ⁱ": "i", "ʲ": "j",
                   "ᵏ": "k", "ˡ": "l", "ᵐ": "m", "ⁿ": "n", "ᵒ": "o", "ᵖ": "p", "ʳ": "r", "ˢ": "s", "ᵗ": "t", "ᵘ": "u",
                   "ᵛ": "v", "ʷ": "w", "ˣ": "x", "ʸ": "y", "ᶻ": "z"}
        for k, v in sub_map.items():
            txt = txt.replace(k, v)
        for k, v in sup_map.items():
            txt = txt.replace(k, v)

        # 常见符号/单位
        txt = re.sub(r"(\d+)\s*[°℃℉]\s*[CKF]?", r"\1°C", txt, flags=re.IGNORECASE)  # 温度统一 → °C（保守）
        txt = re.sub(r"(\d+)\s*(atm|bar|pa|kpa|mpa)", r"\1atm", txt, flags=re.IGNORECASE)  # 压力示意归一

        txt = txt.replace("×", "*").replace("÷", "/").replace("≤", "<=").replace("≥", ">=").replace("≠", "!=").replace("≈", "~")
        txt = txt.replace("→", "->").replace("←", "<-").replace("↔", "<->").replace("⟶", "->").replace("⟵", "<-").replace("⟷", "<->")
        txt = txt.replace("（", "(").replace("）", ")").replace("【", "[").replace("】", "]").replace("｛", "{").replace("｝", "}")
        txt = re.sub(r"\s+", " ", txt).replace(" ", "")  # 清空白
        return txt.lower()

    def _canonical_condition_string(self, keep_dup: bool = True) -> str:
        """将所有【文本条件框】合并为"字符多重集"的顺序无关串。"""
        texts = []
        for idx in self.conditions:
            bb = self.bboxes[idx]
            if bb.is_mol:
                continue
            t = (bb.data.get("text") or "").strip()
            if t:
                texts.append(self._normalize_text(t))
        if not texts:
            return ""

        chars = []
        for t in texts:
            chars.extend(list(t))
        if not keep_dup:
            chars = list(set(chars))
        chars.sort()
        return "".join(chars)

    def _process_role(self, role_data, bboxes_source: List[BBox]) -> List[int]:
        """GT: role_data 中为 int 索引；Pred: role_data 中为 dict，需要临时构造 BBox。"""
        out = []
        for item in role_data:
            bbox = bboxes_source[item] if isinstance(item, int) else BBox(item, self.image_data, xyxy=True, normalized=True)
            self.bboxes.append(bbox)
            out.append(len(self.bboxes) - 1)
        return out

    def schema(self, mol_only: bool = False) -> Tuple[List[int], List[int], List[int]]:
        """根据 mol_only 过滤；可选 filter_recover 兜底防空。"""
        if not mol_only:
            return self.reactants, self.conditions, self.products

        r = [i for i in self.reactants if self.bboxes[i].is_mol]
        c = [i for i in self.conditions if self.bboxes[i].is_mol]
        p = [i for i in self.products if self.bboxes[i].is_mol]

        if self.filter_recover:
            if not r and self.reactants:
                r = self.reactants
            if not p and self.products:
                p = self.products
        return r, c, p

    def is_empty(self, mol_only: bool = False) -> bool:
        r, c, p = self.schema(mol_only)
        return not (r or c or p)

    def compare(self, other: "Reaction", mol_only: bool = False, match_type: str = "hard") -> bool:
        """根据模式（soft/hard/hybrid）比较两条反应是否匹配。"""
        if mol_only or match_type == "soft":
            return self._compare_soft(other)
        if match_type == "hybrid":
            return self._compare_hybrid(other)
        return self._compare_hard(other)

    # ---------- Soft：仅分子，条件忽略 ----------
    def _compare_soft(self, other: "Reaction") -> bool:
        if self.is_empty(True) and other.is_empty(True):
            return True
        if self.is_empty(True) or other.is_empty(True):
            return False

        r1, c1, p1 = self.schema(mol_only=True)
        r2, c2, p2 = other.schema(mol_only=True)

        # 将条件中的分子也视为反应物
        r1 = sorted(set(r1 + c1))
        r2 = sorted(set(r2 + c2))
        p1 = sorted(set(p1))
        p2 = sorted(set(p2))

        if len(r1) != len(r2) or len(p1) != len(p2):
            return False

        if r1:
            m, _, _ = get_bboxes_match([self.bboxes[i] for i in r1], [other.bboxes[i] for i in r2], iou_thres=0.5)
            if (m == -1).any():
                return False
        if p1:
            m, _, _ = get_bboxes_match([self.bboxes[i] for i in p1], [other.bboxes[i] for i in p2], iou_thres=0.5)
            if (m == -1).any():
                return False
        return True

    # ---------- Hybrid：先 IoU，再对未匹配的文本做严格一致 ----------
    def _compare_hybrid(self, other: "Reaction") -> bool:
        if self.is_empty(False) and other.is_empty(False):
            return True
        if self.is_empty(False) or other.is_empty(False):
            return False

        # 1) 反应物/产物逐角色比较
        for role in ("reactants", "products"):
            self_ids = getattr(self, role)
            other_ids = getattr(other, role)
            if len(self_ids) != len(other_ids):
                return False

            s_boxes = [self.bboxes[i] for i in self_ids]
            o_boxes = [other.bboxes[i] for i in other_ids]

            m, _, _ = get_bboxes_match(s_boxes, o_boxes, iou_thres=0.5)
            unmatched_self = [i for i, v in enumerate(m) if v == -1]
            unmatched_other = [i for i in range(len(o_boxes)) if i not in m]

            if unmatched_self and unmatched_other:
                # 用文本严格一致兜底
                cost = np.ones((len(unmatched_self), len(unmatched_other)))
                for i, si in enumerate(unmatched_self):
                    s_txt = s_boxes[si].text
                    for j, oj in enumerate(unmatched_other):
                        o_txt = o_boxes[oj].text
                        if s_txt and o_txt:
                            s_norm = self._normalize_text(s_txt)
                            o_norm = other._normalize_text(o_txt)
                            if s_norm == o_norm:
                                cost[i, j] = 0
                ri, cj = linear_sum_assignment(cost)
                text_ok = sum(1 for i, j in zip(ri, cj) if cost[i, j] == 0)
                total_ok = (m != -1).sum() + text_ok
                if total_ok != len(self_ids):
                    return False
            else:
                if (m != -1).sum() != len(self_ids):
                    return False

        # 2) 条件：分子框做 IoU、文本做"字符多重集"+ 编辑距离阈值
        s_cond_ids, o_cond_ids = self.conditions, other.conditions
        s_cond_mol = [i for i in s_cond_ids if self.bboxes[i].is_mol]
        o_cond_mol = [i for i in o_cond_ids if other.bboxes[i].is_mol]
        if len(s_cond_mol) != len(o_cond_mol):
            return False
        if s_cond_mol:
            m, _, _ = get_bboxes_match([self.bboxes[i] for i in s_cond_mol],
                                       [other.bboxes[i] for i in o_cond_mol], iou_thres=0.5)
            if (m == -1).any():
                return False

        s_str = self._canonical_condition_string(keep_dup=False)
        o_str = other._canonical_condition_string(keep_dup=False)
        if bool(s_str) ^ bool(o_str):
            return False
        if s_str and o_str:
            if normalized_edit_distance(s_str, o_str) > COND_TEXT_EDIT_DIST_THRES:
                return False
        return True

    # ---------- Hard：所有框都必须 IoU 成功 ----------
    def _compare_hard(self, other: "Reaction") -> bool:
        if self.is_empty(False) and other.is_empty(False):
            return True
        if self.is_empty(False) or other.is_empty(False):
            return False

        for role in ("reactants", "products"):
            self_ids = getattr(self, role)
            other_ids = getattr(other, role)
            if len(self_ids) != len(other_ids):
                return False
            m, _, _ = get_bboxes_match([self.bboxes[i] for i in self_ids],
                                       [other.bboxes[i] for i in other_ids], iou_thres=0.5)
            if (m == -1).any():
                return False

        m, _, _ = get_bboxes_match([self.bboxes[i] for i in self.conditions],
                                   [other.bboxes[i] for i in other.conditions], iou_thres=0.5)
        if (m == -1).any():
            return False
        return True


class ReactionImageData:
    """一张图的 GT/Pred 数据与评估。"""

    def __init__(self, gold_data: Dict, pred_data: Dict, image_base_path: str, filter_recover: bool = False):
        g_name = gold_data.get("file_name", "")
        p_name = pred_data.get("file_name", "")
        self.file_name = g_name
        if g_name != p_name:
            print("Warning: Filename mismatch:\n  Gold:", g_name, "\n  Pred:", p_name)

        image_path = os.path.join(image_base_path, self.file_name)
        img = cv2.imread(image_path)
        if img is not None:
            self.height, self.width = img.shape[:2]
        else:
            self.width = gold_data.get("width", 1000)
            self.height = gold_data.get("height", 1000)
            print(f"Warning: Could not load image {image_path}. Using dimensions from JSON.")

        self.gold_bboxes = [BBox(b, self, xyxy=False, normalized=False) for b in gold_data.get("bboxes", [])]
        self.pred_bboxes = [BBox(b, self, xyxy=False, normalized=False) for b in pred_data.get("bboxes", [])]
        self.gold_reactions = [Reaction(r, self.gold_bboxes, self, filter_recover=filter_recover) for r in gold_data.get("reactions", [])]
        self.pred_reactions = [Reaction(r, self.pred_bboxes, self, filter_recover=filter_recover) for r in pred_data.get("reactions", [])]

    def evaluate(self, mol_only: bool = False, match_type: str = "hard") -> Dict:
        """对当前图片计算 TP/FP/FN 与匹配对。"""
        golds = [r for r in self.gold_reactions if not r.is_empty(mol_only)]
        preds = [r for r in self.pred_reactions if not r.is_empty(mol_only)]

        if not golds and not preds:
            # 如果原始反应非空但筛后为空，标记 ignored，便于上层统计
            if self.gold_reactions or self.pred_reactions:
                return {"ignored": True, "tp": 0, "fp": 0, "fn": 0, "gold_total": 0, "pred_total": 0, "matches": []}
            return {"tp": 0, "fp": 0, "fn": 0, "gold_total": 0, "pred_total": 0, "matches": []}

        cost = np.ones((len(golds), len(preds)))
        for i, g in enumerate(golds):
            for j, p in enumerate(preds):
                if g.compare(p, mol_only=mol_only, match_type=match_type):
                    cost[i, j] = 0

        row, col = linear_sum_assignment(cost)
        matches = [(r, c) for r, c in zip(row, col) if cost[r, c] == 0]
        tp = len(matches)
        fp = len(preds) - tp
        fn = len(golds) - tp
        return {"tp": tp, "fp": fp, "fn": fn, "gold_total": len(golds), "pred_total": len(preds), "matches": matches}


# =========================
# 工具函数
# =========================
def get_iou(bb1: BBox, bb2: BBox) -> float:
    x_left = max(bb1.x1, bb2.x1)
    y_top = max(bb1.y1, bb2.y1)
    x_right = min(bb1.x2, bb2.x2)
    y_bottom = min(bb1.y2, bb2.y2)
    if x_right < x_left or y_bottom < y_top:
        return 0.0
    inter = (x_right - x_left) * (y_bottom - y_top)
    a1 = (bb1.x2 - bb1.x1) * (bb1.y2 - bb1.y1)
    a2 = (bb2.x2 - bb2.x1) * (bb2.y2 - bb2.y1)
    denom = a1 + a2 - inter
    return 0.0 if denom <= 0 else inter / denom


def get_bboxes_match(bboxes1: List[BBox], bboxes2: List[BBox], iou_thres: float = 0.5):
    """用匈牙利在最大 IoU 上做一一匹配，返回 match1（长度=len(bboxes1)）。"""
    if not bboxes1 or not bboxes2:
        n1, n2 = len(bboxes1), len(bboxes2)
        return np.full(n1, -1, dtype=int), np.full(n2, -1, dtype=int), np.zeros((n1, n2))

    scores = np.zeros((len(bboxes1), len(bboxes2)))
    for i, b1 in enumerate(bboxes1):
        for j, b2 in enumerate(bboxes2):
            scores[i, j] = get_iou(b1, b2)

    match1 = np.full(len(bboxes1), -1, dtype=int)
    row, col = linear_sum_assignment(-scores)
    for r, c in zip(row, col):
        if scores[r, c] >= iou_thres:
            match1[r] = c

    match2 = np.full(len(bboxes2), -1, dtype=int)
    for i, m in enumerate(match1):
        if m != -1:
            match2[m] = i
    return match1, match2, scores


def compute_metrics(tp: int, gold_total: int, pred_total: int) -> Dict[str, float]:
    precision = tp / max(pred_total, 1)
    recall = tp / max(gold_total, 1)
    f1 = 2 * precision * recall / max(precision + recall, 1e-6)
    return {"precision": precision, "recall": recall, "f1": f1}


# =========================
# 可视化（单对反应）
# =========================
def create_reaction_level_comparison(
    img_rgb, pair, output_dir: str, file_name: str, item_idx: int, pair_idx: int, *, mol_only=False, match_type="hard"
) -> str:
    """根据匹配模式生成可视化（soft：1x2；hard/hybrid：2x2 带文本）。"""
    if match_type == "soft":
        nrows, ncols, hspace = 1, 2, 0.02
    else:
        nrows, ncols, hspace = 2, 2, 0.20

    fig, axes = plt.subplots(
        nrows,
        ncols,
        figsize=(20, 9 if match_type == "soft" else 12),
        squeeze=False,
        gridspec_kw={"height_ratios": [1, 0.4] if match_type != "soft" else None, "hspace": hspace},
    )
    role_colors = {"reactants": "red", "conditions": "blue", "products": "green"}

    # 左上：GT
    ax_gt = axes[0, 0]
    ax_gt.imshow(img_rgb)
    ax_gt.set_axis_off()
    ax_gt.set_title(("Ground Truth" if pair["gt"] else "Closest GT (None)") + f" – {pair['status']}", fontsize=12)

    if pair["gt"]:
        for role, ids in [("reactants", pair["gt"].reactants), ("conditions", pair["gt"].conditions), ("products", pair["gt"].products)]:
            for bid in ids:
                x, y, w, h = pair["gt"].bboxes[bid].to_pixel()
                ax_gt.add_patch(patches.Rectangle((x, y), w, h, linewidth=2, edgecolor=role_colors.get(role, "gray"), facecolor="none"))

    # 右上：Pred
    ax_pred = axes[0, 1]
    ax_pred.imshow(img_rgb)
    ax_pred.set_axis_off()
    title_color = "green" if "TP" in pair["status"] else ("red" if "FP" in pair["status"] else "gray")
    ax_pred.set_title("Prediction – " + pair["status"], fontsize=12, color=title_color)

    if pair["pred"]:
        all_gt = pair["gt"].bboxes if pair["gt"] else []
        all_pred = pair["pred"].bboxes
        _, _, iou_matrix = get_bboxes_match(all_gt, all_pred)
        for pid, bb in enumerate(all_pred):
            x, y, w, h = bb.to_pixel()
            max_iou = np.max(iou_matrix[:, pid]) if iou_matrix.size else 0.0
            is_loc_ok = max_iou >= 0.5
            ax_pred.add_patch(
                patches.Rectangle(
                    (x, y),
                    w,
                    h,
                    linewidth=2,
                    edgecolor=role_colors.get(
                        "reactants" if pid in pair["pred"].reactants else "conditions" if pid in pair["pred"].conditions else "products" if pid in pair["pred"].products else "gray"
                    ),
                    linestyle="-" if is_loc_ok else "--",
                    facecolor="none",
                )
            )
            ax_pred.text(x, y - 5, f"IoU:{max_iou:.2f}", color="darkgreen" if is_loc_ok else "darkred", fontsize=7)

    # 下排文本（仅 hard/hybrid）
    if match_type != "soft":
        ax_gt_txt, ax_pred_txt = axes[1, 0], axes[1, 1]
        for a in (ax_gt_txt, ax_pred_txt):
            a.axis("off")

        def role_texts(react: Optional[Reaction]) -> Dict[str, List[str]]:
            out = {"reactants": [], "conditions": [], "products": []}
            if not react:
                return out
            for role, ids in [("reactants", react.reactants), ("conditions", react.conditions), ("products", react.products)]:
                for bid in ids:
                    bb = react.bboxes[bid]
                    if not bb.is_mol and bb.text:
                        out[role].append(bb.text.strip())
            return out

        def draw_text(ax, txt_dict, label):
            lines = []
            for role in ("reactants", "conditions", "products"):
                ts = txt_dict[role]
                if ts:
                    lines.append(f"$\\bf{{{role[:3].upper()}}}$: " + ", ".join(ts))
            ax.text(0, 0.98, "\n".join(lines) if lines else "(none)", va="top", ha="left", fontsize=10, wrap=True)
            ax.set_title(label, fontsize=11, pad=2)

        draw_text(ax_gt_txt, role_texts(pair["gt"]), "GT Texts")
        draw_text(ax_pred_txt, role_texts(pair["pred"]), "Pred Texts")

    # 保存
    mode_label = "Soft" if match_type == "soft" else ("Hybrid" if match_type == "hybrid" else "Hard")
    fig.suptitle(f"{file_name} – Pair {pair_idx+1} ({mode_label})", fontsize=15)
    plt.tight_layout(rect=[0, 0, 1, 0.94])

    vis_dir = os.path.join(output_dir, "visualizations")
    os.makedirs(vis_dir, exist_ok=True)
    safe_name = file_name.replace("/", "_").replace("\\", "_")
    out_path = os.path.join(vis_dir, f"{item_idx:04d}_{pair_idx:02d}_{mode_label.lower()}_{safe_name}.png")
    plt.savefig(out_path, dpi=110)
    plt.close(fig)
    return out_path


# =========================
# 汇总评估（Micro）
# =========================
def run_overall_evaluation(gold_images, pred_images, image_base_path: str, filter_recover: bool = False) -> None:
    print("\n" + "=" * 30 + " Overall & Grouped Evaluation Summary " + "=" * 30)
    eval_types = {
        "Hard Match": {"mol_only": False, "match_type": "hard"},
        "Soft Match": {"mol_only": True, "match_type": "soft"},
        "Hybrid Match": {"mol_only": False, "match_type": "hybrid"},
    }

    for name, params in eval_types.items():
        print(f"\n--- {name} ---")

        all_results = []
        group_results = defaultdict(list)
        group_counts = defaultdict(int)
        group_ignored = defaultdict(int)

        gold_by_name = {g["file_name"]: g for g in gold_images}
        pred_by_name = {p["file_name"]: p for p in pred_images}
        for fn in sorted(set(gold_by_name) | set(pred_by_name)):
            g = gold_by_name.get(fn)
            p = pred_by_name.get(fn)
            if not g or not p:
                continue

            data = ReactionImageData(g, p, image_base_path, filter_recover=filter_recover)
            res = data.evaluate(**params)
            grp = g.get("diagram_type", "unknown")
            if res.get("ignored"):
                group_ignored[grp] += 1
                continue
            all_results.append(res)
            group_results[grp].append(res)
            group_counts[grp] += 1

        tp = sum(r["tp"] for r in all_results)
        gold_tot = sum(r["gold_total"] for r in all_results)
        pred_tot = sum(r["pred_total"] for r in all_results)
        metrics = compute_metrics(tp, gold_tot, pred_tot)
        sum_fp = sum(r["fp"] for r in all_results)
        sum_fn = sum(r["fn"] for r in all_results)
        print(f"  {'Overall (Micro)':<15} | TP:{tp} FP:{sum_fp} FN:{sum_fn} | Gold:{gold_tot} Pred:{pred_tot} | P: {metrics['precision']:.4f}, R: {metrics['recall']:.4f}, F1: {metrics['f1']:.4f} ({len(all_results)} images, {sum(group_ignored.values())} ignored)")

        for grp in sorted(set(group_counts) | set(group_ignored)):
            lst = group_results.get(grp, [])
            tp_g = sum(r["tp"] for r in lst)
            gtot = sum(r["gold_total"] for r in lst)
            ptot = sum(r["pred_total"] for r in lst)
            sum_fp_g = sum(r["fp"] for r in lst)
            sum_fn_g = sum(r["fn"] for r in lst)
            m = compute_metrics(tp_g, gtot, ptot)
            print(f"  {grp:<15} | TP:{tp_g} FP:{sum_fp_g} FN:{sum_fn_g} | Gold:{gtot} Pred:{ptot} | P: {m['precision']:.4f}, R: {m['recall']:.4f}, F1: {m['f1']:.4f} ({group_counts.get(grp, 0)} images, {group_ignored.get(grp, 0)} ignored)")

    print("=" * 100 + "\n")


# =========================
# XLSX（逐项）报告
# =========================
def save_results_to_xlsx(report_items: List[Dict], output_dir: str, report_name_suffix: str, ignored_count: int, vis_count: int) -> None:
    if not report_items:
        print("\nNo items to include in the XLSX report.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Report"

    headers = [
        "Item Index",
        "File Name",
        "GT Reactions",
        "Pred Reactions",
        "Hard_TP",
        "Hard_FP",
        "Hard_FN",
        "Hard_FP+FN",
        "Hard_Precision",
        "Hard_Recall",
        "Hard_F1",
        "Soft_TP",
        "Soft_FP",
        "Soft_FN",
        "Soft_FP+FN",
        "Soft_Precision",
        "Soft_Recall",
        "Soft_F1",
        "Hybrid_TP",
        "Hybrid_FP",
        "Hybrid_FN",
        "Hybrid_FP+FN",
        "Hybrid_Precision",
        "Hybrid_Recall",
        "Hybrid_F1",
        "Visualization Path",
    ]
    ws.append(headers)

    total_gt, total_pred, total_hard_tp, total_soft_tp, total_soft_pred, total_hybrid_tp, total_hybrid_pred = 0, 0, 0, 0, 0, 0, 0
    for r in report_items:
        hm, sm = r["hard_match"], r["soft_match"]
        hm_m = compute_metrics(hm["tp"], hm["gold_total"], hm["pred_total"])
        sm_m = compute_metrics(sm["tp"], sm["gold_total"], sm["pred_total"]) if sm else {"precision": 0, "recall": 0, "f1": 0}
        hy = r.get("hybrid_match")
        hy_m = compute_metrics(hy["tp"], hy["gold_total"], hy["pred_total"]) if hy else {"precision": 0, "recall": 0, "f1": 0}
        total_gt += hm["gold_total"]
        total_pred += hm["pred_total"]
        total_hard_tp += hm["tp"]
        if sm:
            total_soft_tp += sm["tp"]
            total_soft_pred += sm["pred_total"]
        if hy:
            total_hybrid_tp += hy["tp"]
            total_hybrid_pred += hy["pred_total"]

        vis_path = r.get("visualization_path") or ""
        if isinstance(vis_path, list):
            vis_path = vis_path[0] if vis_path else ""

        ws.append(
            [
                r["item_index"],
                r["file_name"],
                hm["gold_total"],
                hm["pred_total"],
                hm["tp"],
                hm["fp"],
                hm["fn"],
                hm["fp"] + hm["fn"],
                f"{hm_m['precision']:.3f}",
                f"{hm_m['recall']:.3f}",
                f"{hm_m['f1']:.3f}",
                sm.get("tp") if sm else "N/A",
                sm.get("fp") if sm else "N/A",
                sm.get("fn") if sm else "N/A",
                (sm.get("fp", 0) + sm.get("fn", 0)) if sm else "N/A",
                f"{sm_m['precision']:.3f}",
                f"{sm_m['recall']:.3f}",
                f"{sm_m['f1']:.3f}",
                hy.get("tp") if hy else "N/A",
                hy.get("fp") if hy else "N/A",
                hy.get("fn") if hy else "N/A",
                (hy.get("fp", 0) + hy.get("fn", 0)) if hy else "N/A",
                f"{hy_m['precision']:.3f}",
                f"{hy_m['recall']:.3f}",
                f"{hy_m['f1']:.3f}",
                vis_path,
            ]
        )

    overall_hard_precision = total_hard_tp / max(total_pred, 1)
    overall_hard_recall = total_hard_tp / max(total_gt, 1)
    overall_soft_precision = total_soft_tp / max(total_soft_pred, 1)
    overall_soft_recall = total_soft_tp / max(total_gt, 1)
    overall_hybrid_precision = total_hybrid_tp / max(total_hybrid_pred, 1)
    overall_hybrid_recall = total_hybrid_tp / max(total_gt, 1)

    ws.append([])
    ws.append(["Summary"])
    ws.append(
        [
            "Total GT Reactions",
            "Total Pred Reactions",
            "Total Hard_TP",
            "Total Soft_TP",
            "Total Hybrid_TP",
            "Overall Hard Precision",
            "Overall Hard Recall",
            "Overall Soft Precision",
            "Overall Soft Recall",
            "Overall Hybrid Precision",
            "Overall Hybrid Recall",
        ]
    )
    ws.append(
        [
            total_gt,
            total_pred,
            total_hard_tp,
            total_soft_tp,
            total_hybrid_tp,
            f"{overall_hard_precision:.4f}",
            f"{overall_hard_recall:.4f}",
            f"{overall_soft_precision:.4f}",
            f"{overall_soft_recall:.4f}",
            f"{overall_hybrid_precision:.4f}",
            f"{overall_hybrid_recall:.4f}",
        ]
    )

    for i, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions[get_column_letter(len(headers))].width = 60

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = os.path.join(output_dir, f"{report_name_suffix}_{timestamp}.xlsx")
    wb.save(xlsx_path)
    print(f"\nDetailed XLSX report saved to: {xlsx_path}")
    if ignored_count is not None:
        print(f"Ignored cases (GT and Pred both empty): {ignored_count}")
    if vis_count > 0:
        print(f"Bad case visualizations saved in: {os.path.join(output_dir, 'bad_cases_visualizations')}")


def run_itemized_evaluation(
    gold_images,
    pred_images,
    output_dir: str,
    image_base_path: str,
    vis_range: Optional[Tuple[int, int]],
    limit_vis: int = -1,
    generate_report: bool = True,
    report_only_visualized: bool = False,
    filter_recover: bool = False,
) -> None:
    """逐文件计算 Hard/Soft，并可选输出 XLSX。可视化在 Markdown 里完成，这里不画图。"""
    print("Starting item-by-item evaluation...")
    results = []
    vis_count = 0
    ignored_count = 0

    gold_by_name = {g["file_name"]: g for g in gold_images}
    pred_by_name = {p["file_name"]: p for p in pred_images}

    for idx, fn in enumerate(sorted(set(gold_by_name) | set(pred_by_name))):
        g = gold_by_name.get(fn)
        p = pred_by_name.get(fn)
        if not g or not p:
            continue

        data = ReactionImageData(g, p, image_base_path, filter_recover=filter_recover)
        hard_match = data.evaluate(mol_only=False, match_type="hard")
        if hard_match.get("ignored"):
            ignored_count += 1
            continue
        soft_match = data.evaluate(mol_only=True, match_type="soft") if generate_report else None
        hybrid_match = data.evaluate(mol_only=False, match_type="hybrid") if generate_report else None
        results.append({"item_index": idx, "file_name": fn, "hard_match": hard_match, "soft_match": soft_match, "hybrid_match": hybrid_match, "visualization_path": None})

        # 只用于统计（是否在指定 FP+FN 区间内有坏例）
        fp_fn = hard_match["fp"] + hard_match["fn"]
        if vis_range:
            if vis_range[0] <= fp_fn <= vis_range[1]:
                vis_count += 1
        else:
            if fp_fn > 0:
                vis_count += 1

    if not generate_report:
        print("\nNo XLSX generated (generate_report=False).")
        return

    report_items = results if not report_only_visualized else [r for r in results if r["visualization_path"]]
    if not report_items:
        print("\nNo items to include in the report.")
        return

    save_results_to_xlsx(report_items, output_dir, "comprehensive_report", ignored_count, vis_count)


# =========================
# Markdown（按模式）可视化
# =========================
def generate_md_for_setting(
    gold_images,
    pred_images,
    image_base_path: str,
    output_dir: str,
    vis_range: Optional[Tuple[int, int]],
    limit_vis: int,
    *,
    mol_only: bool = False,
    match_type: str = "hard",
    filter_recover: bool = False,
) -> None:
    """筛选"有误差的样本"，输出 Markdown + 每个样本的配图对比。"""
    mode_tag = match_type
    print(f"\n[Markdown-{mode_tag}] collecting cases ...")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    report_dir = os.path.join(output_dir, f"md_{mode_tag}_{timestamp}")
    os.makedirs(report_dir, exist_ok=True)

    results, ignored = [], 0
    gold_by_name = {g["file_name"]: g for g in gold_images}
    pred_by_name = {p["file_name"]: p for p in pred_images}

    for idx, fn in enumerate(sorted(set(gold_by_name) | set(pred_by_name))):
        g = gold_by_name.get(fn)
        p = pred_by_name.get(fn)
        if not g or not p:
            continue

        data = ReactionImageData(g, p, image_base_path, filter_recover=filter_recover)
        current_eval = data.evaluate(mol_only=mol_only, match_type=match_type)
        hard_eval = data.evaluate(mol_only=False, match_type="hard")
        soft_eval = data.evaluate(mol_only=True, match_type="soft")
        hybrid_eval = data.evaluate(mol_only=False, match_type="hybrid")

        if hard_eval.get("ignored"):
            ignored += 1
            continue

        results.append(
            {
                "item_index": idx,
                "file_name": fn,
                "hard_match": hard_eval,
                "soft_match": soft_eval,
                "hybrid_match": hybrid_eval,
                "current_eval": current_eval,
                "data": data,
            }
        )

    # 按当前模式过滤"有误差"的样本并排序
    key = {"soft": "soft_match", "hard": "hard_match", "hybrid": "hybrid_match"}.get(match_type, "current_eval")
    results = [r for r in results if r[key]["fp"] + r[key]["fn"] > 0]
    results.sort(key=lambda r: r[key]["fp"] + r[key]["fn"], reverse=True)
    if limit_vis > 0:
        results = results[:limit_vis]

    vis_count = 0
    md_path = os.path.join(report_dir, f"{mode_tag}_match_errors.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(f"# {mode_tag.capitalize()}-Match Error Report\n\n")
        f.write(f"Generated on: {timestamp}\n\n")
        f.write(f"Total Visualized: {len(results)} cases\n\n---\n\n")

        for res in results:
            idx, fn, data = res["item_index"], res["file_name"], res["data"]
            hard_eval, soft_eval, hybrid_eval, cur_eval = res["hard_match"], res["soft_match"], res["hybrid_match"], res["current_eval"]

            hm = compute_metrics(hard_eval["tp"], hard_eval["gold_total"], hard_eval["pred_total"])
            sm = compute_metrics(soft_eval["tp"], soft_eval["gold_total"], soft_eval["pred_total"])
            hy = compute_metrics(hybrid_eval["tp"], hybrid_eval["gold_total"], hybrid_eval["pred_total"])

            f.write(f"## Case: `{fn}` (Index: {idx})\n\n")
            f.write("**Overall Metrics for this Image:**\n")
            f.write("| Match Type     | TP | FP | FN | Precision | Recall | F1-Score |\n")
            f.write("|----------------|----|----|----|-----------|--------|----------|\n")
            f.write(f"| **Hard**       | {hard_eval['tp']} | {hard_eval['fp']} | {hard_eval['fn']} | {hm['precision']:.3f} | {hm['recall']:.3f} | {hm['f1']:.3f} |\n")
            f.write(f"| **Soft**       | {soft_eval['tp']} | {soft_eval['fp']} | {soft_eval['fn']} | {sm['precision']:.3f} | {sm['recall']:.3f} | {sm['f1']:.3f} |\n")
            f.write(f"| **Hybrid**     | {hybrid_eval['tp']} | {hybrid_eval['fp']} | {hybrid_eval['fn']} | {hy['precision']:.3f} | {hy['recall']:.3f} | {hy['f1']:.3f} |\n\n")

            # 准备配对（TP、FP、FN）
            gt = [r for r in data.gold_reactions if not r.is_empty(mol_only)]
            pr = [r for r in data.pred_reactions if not r.is_empty(mol_only)]
            match = cur_eval["matches"]
            matched_gt = {m[0] for m in match}
            matched_pr = {m[1] for m in match}
            un_gt = [i for i in range(len(gt)) if i not in matched_gt]
            un_pr = [i for i in range(len(pr)) if i not in matched_pr]

            # 将 FP 映射到最相近的 GT（仅用于配图时的"参考"）
            fp_best = {}
            if un_gt and un_pr:
                for pi in un_pr:
                    pred_rxn = pr[pi]
                    best_sum, best_idx = 0, None
                    for gi in un_gt:
                        gt_rxn = gt[gi]
                        m, _, ious = get_bboxes_match(gt_rxn.bboxes, pred_rxn.bboxes)
                        tot = sum(ious[g, p] for g, p in enumerate(m) if p != -1)
                        if tot > best_sum:
                            best_sum, best_idx = tot, gi
                    if best_idx is not None:
                        fp_best[pi] = best_idx

            display_pairs = []
            for gi, pj in match:
                m, _, ious = get_bboxes_match(gt[gi].bboxes, pr[pj].bboxes)
                cnt = sum(1 for v in m if v != -1)
                avg_iou = (sum(ious[g, p] for g, p in enumerate(m) if p != -1) / cnt) if cnt else 0
                status = f"Matched (TP) - Avg IoU: {avg_iou:.2f}" if avg_iou >= 0.5 else f"Poor Match - Avg IoU: {avg_iou:.2f}"
                display_pairs.append({"gt": gt[gi], "pred": pr[pj], "status": status})
            for pi in un_pr:
                display_pairs.append({"gt": gt[fp_best[pi]] if pi in fp_best else None, "pred": pr[pi], "status": "Wrong Pred (FP)"})
            for gi in un_gt:
                if gi not in fp_best.values():
                    display_pairs.append({"gt": gt[gi], "pred": None, "status": "Missed GT (FN)"})

            img_path = os.path.join(image_base_path, fn)
            img = cv2.imread(img_path)
            if img is None:
                print(f"  -> Skipped: cannot read image {img_path}")
                continue
            img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

            for k, pair in enumerate(display_pairs):
                vis_path = create_reaction_level_comparison(img_rgb, pair, report_dir, fn, idx, k, mol_only=mol_only, match_type=match_type)
                if vis_path:
                    rel_path = os.path.relpath(vis_path, report_dir).replace("\\", "/")
                    f.write(f"**{pair['status']}**\n\n")
                    f.write(f"![Visualization]({rel_path})\n\n")
                    vis_count += 1
            f.write("---\n\n")

    print(f"\n✅ {mode_tag.capitalize()} Match Report saved to: {md_path}")
    print(f"   Total visualizations: {vis_count}")
    if ignored:
        print(f"   Ignored empty images: {ignored}")


# =========================
# 汇总到 Excel（Hard/Soft/Hybrid × Overall/single/multiple/tree/graph）
# =========================
def save_match_results_to_excel(gold_images, pred_images, image_base_path: str, output_dir: str, filter_recover: bool = False) -> None:
    print("\n" + "=" * 30 + " Exporting Match Results to Excel " + "=" * 30)
    eval_types = {"Hard": {"mol_only": False, "match_type": "hard"}, "Soft": {"mol_only": True, "match_type": "soft"}, "Hybrid": {"mol_only": False, "match_type": "hybrid"}}
    group_types = ["Overall", "single", "multiple", "tree", "graph"]

    wb = openpyxl.Workbook()
    gold_by_name = {g["file_name"]: g for g in gold_images}
    pred_by_name = {p["file_name"]: p for p in pred_images}

    for k, (match_name, params) in enumerate(eval_types.items()):
        ws = wb.active if k == 0 else wb.create_sheet(match_name)
        ws.title = match_name

        headers = ["", "TP", "FP", "FN", "Gold", "Pred", "P", "R", "F1"]
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.fill = openpyxl.styles.PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

        for r, grp in enumerate(group_types, 2):
            cell = ws.cell(row=r, column=1, value=grp)
            cell.font = openpyxl.styles.Font(bold=True)

        all_res = []
        group_res = defaultdict(list)
        group_cnt = defaultdict(int)
        group_ignored = defaultdict(int)

        for fn in sorted(set(gold_by_name) | set(pred_by_name)):
            g = gold_by_name.get(fn)
            p = pred_by_name.get(fn)
            if not g or not p:
                continue
            data = ReactionImageData(g, p, image_base_path, filter_recover=filter_recover)
            res = data.evaluate(**params)
            grp = g.get("diagram_type", "unknown")
            if res.get("ignored"):
                group_ignored[grp] += 1
                continue
            all_res.append(res)
            group_res[grp].append(res)
            group_cnt[grp] += 1

        tp = sum(r["tp"] for r in all_res)
        gt = sum(r["gold_total"] for r in all_res)
        pd = sum(r["pred_total"] for r in all_res)
        fp = sum(r["fp"] for r in all_res)
        fn = sum(r["fn"] for r in all_res)
        overall = compute_metrics(tp, gt, pd)
        # Fill Overall row
        ws.cell(row=2, column=2, value=tp)
        ws.cell(row=2, column=3, value=fp)
        ws.cell(row=2, column=4, value=fn)
        ws.cell(row=2, column=5, value=gt)
        ws.cell(row=2, column=6, value=pd)
        ws.cell(row=2, column=7, value=round(overall["precision"] * 100, 2))
        ws.cell(row=2, column=8, value=round(overall["recall"] * 100, 2))
        ws.cell(row=2, column=9, value=round(overall["f1"] * 100, 2))
        print(f"  {match_name:<6} Overall: TP={tp}, FP={fp}, FN={fn}, Gold={gt}, Pred={pd} | P={overall['precision']*100:.2f}%, R={overall['recall']*100:.2f}%, F1={overall['f1']*100:.2f}%")

        for grp in ["single", "multiple", "tree", "graph"]:
            lst = group_res.get(grp, [])
            row_idx = group_types.index(grp) + 2
            if lst:
                tp_g = sum(r_["tp"] for r_ in lst)
                gtot = sum(r_["gold_total"] for r_ in lst)
                ptot = sum(r_["pred_total"] for r_ in lst)
                fp_g = sum(r_["fp"] for r_ in lst)
                fn_g = sum(r_["fn"] for r_ in lst)
                m = compute_metrics(tp_g, gtot, ptot)
                ws.cell(row=row_idx, column=2, value=tp_g)
                ws.cell(row=row_idx, column=3, value=fp_g)
                ws.cell(row=row_idx, column=4, value=fn_g)
                ws.cell(row=row_idx, column=5, value=gtot)
                ws.cell(row=row_idx, column=6, value=ptot)
                ws.cell(row=row_idx, column=7, value=round(m["precision"] * 100, 2))
                ws.cell(row=row_idx, column=8, value=round(m["recall"] * 100, 2))
                ws.cell(row=row_idx, column=9, value=round(m["f1"] * 100, 2))
                print(f"  {match_name:<6} {grp:<8}: TP={tp_g}, FP={fp_g}, FN={fn_g}, Gold={gtot}, Pred={ptot} | P={m['precision']*100:.2f}%, R={m['recall']*100:.2f}%, F1={m['f1']*100:.2f}% ({len(lst)} images)")
            else:
                for col in range(2, 10):
                    ws.cell(row=row_idx, column=col, value="")
                print(f"  {match_name:<6} {grp:<8}: No data")

        for col in range(1, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"match_results_{timestamp}.xlsx")
    wb.save(excel_path)
    print(f"\n✅ Match results Excel file saved to: {excel_path}")
    print("=" * 100 + "\n")


# =========================
# 每图F1排序导出（Soft/Hybrid）
# =========================

def save_per_image_f1_ranking_to_csv(
    gold_images,
    pred_images,
    image_base_path: str,
    output_dir: str,
    filter_recover: bool = False,
) -> None:
    print("\n" + "=" * 30 + " Exporting Per-Image F1 Rankings (CSV) " + "=" * 30)

    eval_types = {
        "Soft": {"mol_only": True, "match_type": "soft"},
        "Hybrid": {"mol_only": False, "match_type": "hybrid"},
    }

    gold_by_name = {g["file_name"]: g for g in gold_images}
    pred_by_name = {p["file_name"]: p for p in pred_images}

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    for sheet_name, params in eval_types.items():
        rows = []
        for fn in sorted(set(gold_by_name) | set(pred_by_name)):
            g = gold_by_name.get(fn)
            p = pred_by_name.get(fn)
            if not g or not p:
                continue

            data = ReactionImageData(g, p, image_base_path, filter_recover=filter_recover)
            res = data.evaluate(**params)
            if res.get("ignored"):
                continue

            m = compute_metrics(res["tp"], res["gold_total"], res["pred_total"])
            rows.append((fn, res["gold_total"], m["f1"]))

        rows.sort(key=lambda x: (x[2], x[0]))

        csv_path = os.path.join(output_dir, f"per_image_f1_ranking_{sheet_name.lower()}_{timestamp}.csv")
        with open(csv_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["File Name", "GT Count", "F1"])
            for fn, gt_cnt, f1 in rows:
                writer.writerow([fn, gt_cnt, f"{f1:.6f}"])
        print(f"  ✅ {sheet_name}: {csv_path}")

    print("=" * 100 + "\n")


# =========================
# 问题图片收集
# =========================
def collect_problematic_images(
    gold_images: List[Dict], 
    pred_images: List[Dict], 
    image_base_path: str, 
    output_dir: str, 
    filter_recover: bool = False
) -> None:
    """收集有问题的图片名（有错误预测或有漏预测的图片）并保存到txt文件，按匹配类型分类统计"""
    print("\n" + "=" * 30 + " Collecting Problematic Images by Match Type " + "=" * 30)
    
    # 按匹配类型分类收集问题图片
    match_types = {
        "hard": {"mol_only": False, "match_type": "hard"},
        "soft": {"mol_only": True, "match_type": "soft"},
        "hybrid": {"mol_only": False, "match_type": "hybrid"}
    }
    
    problematic_by_type = {match_type: set() for match_type in match_types.keys()}
    all_problematic = set()
    
    gold_by_name = {g["file_name"]: g for g in gold_images}
    pred_by_name = {p["file_name"]: p for p in pred_images}
    
    for fn in sorted(set(gold_by_name) | set(pred_by_name)):
        g = gold_by_name.get(fn)
        p = pred_by_name.get(fn)
        if not g or not p:
            continue
            
        data = ReactionImageData(g, p, image_base_path, filter_recover=filter_recover)
        
        # 检查每种匹配模式是否有问题
        for match_type, params in match_types.items():
            eval_result = data.evaluate(**params)
            
            # 如果有FP或FN，就认为是该类型的问题图片
            if eval_result.get("fp", 0) > 0 or eval_result.get("fn", 0) > 0:
                problematic_by_type[match_type].add(fn)
                all_problematic.add(fn)
    
    # 保存到txt文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    
    # 1. 保存所有问题图片
    all_txt_path = os.path.join(output_dir, f"problematic_images_all_{timestamp}.txt")
    with open(all_txt_path, "w", encoding="utf-8") as f:
        for img_name in sorted(all_problematic):
            f.write(f"{img_name}\n")
    
    # 2. 按类型分别保存
    for match_type, problematic_set in problematic_by_type.items():
        if problematic_set:  # 只保存非空的类型
            type_txt_path = os.path.join(output_dir, f"problematic_images_{match_type}_{timestamp}.txt")
            with open(type_txt_path, "w", encoding="utf-8") as f:
                for img_name in sorted(problematic_set):
                    f.write(f"{img_name}\n")
    
    # 3. 保存交叉分析结果
    cross_analysis_path = os.path.join(output_dir, f"problematic_images_analysis_{timestamp}.txt")
    with open(cross_analysis_path, "w", encoding="utf-8") as f:
        f.write("Problematic Images Analysis Report\n")
        f.write("=" * 50 + "\n\n")
        
        # 总体统计
        f.write(f"Total problematic images: {len(all_problematic)}\n")
        f.write(f"Percentage of total: {len(all_problematic)/len(gold_by_name)*100:.2f}%\n\n")
        
        # 各类型统计
        f.write("Breakdown by match type:\n")
        f.write("-" * 30 + "\n")
        for match_type, problematic_set in problematic_by_type.items():
            f.write(f"{match_type.upper():<8}: {len(problematic_set):>4} images ({len(problematic_set)/len(gold_by_name)*100:>5.2f}%)\n")
        
        # 交叉分析
        f.write(f"\nCross-analysis:\n")
        f.write("-" * 30 + "\n")
        
        # 只在hard中出错
        only_hard = problematic_by_type["hard"] - problematic_by_type["soft"] - problematic_by_type["hybrid"]
        f.write(f"Only in HARD:     {len(only_hard):>4} images\n")
        
        # 只在soft中出错
        only_soft = problematic_by_type["soft"] - problematic_by_type["hard"] - problematic_by_type["hybrid"]
        f.write(f"Only in SOFT:     {len(only_soft):>4} images\n")
        
        # 只在hybrid中出错
        only_hybrid = problematic_by_type["hybrid"] - problematic_by_type["hard"] - problematic_by_type["soft"]
        f.write(f"Only in HYBRID:   {len(only_hybrid):>4} images\n")
        
        # 在hard和soft中都出错
        hard_soft = problematic_by_type["hard"] & problematic_by_type["soft"]
        f.write(f"HARD + SOFT:      {len(hard_soft):>4} images\n")
        
        # 在hard和hybrid中都出错
        hard_hybrid = problematic_by_type["hard"] & problematic_by_type["hybrid"]
        f.write(f"HARD + HYBRID:    {len(hard_hybrid):>4} images\n")
        
        # 在soft和hybrid中都出错
        soft_hybrid = problematic_by_type["soft"] & problematic_by_type["hybrid"]
        f.write(f"SOFT + HYBRID:    {len(soft_hybrid):>4} images\n")
        
        # 在所有三种类型中都出错
        all_three = problematic_by_type["hard"] & problematic_by_type["soft"] & problematic_by_type["hybrid"]
        f.write(f"ALL THREE:        {len(all_three):>4} images\n")
        
        # 详细列表
        f.write(f"\nDetailed lists:\n")
        f.write("-" * 30 + "\n")
        
        for match_type, problematic_set in problematic_by_type.items():
            if problematic_set:
                f.write(f"\n{match_type.upper()} match problems ({len(problematic_set)} images):\n")
                for i, img_name in enumerate(sorted(problematic_set), 1):
                    f.write(f"  {i:3d}. {img_name}\n")
    
    # 输出到控制台
    print(f"✅ Found {len(all_problematic)} total problematic images")
    print(f"✅ Analysis report saved to: {cross_analysis_path}")
    print(f"✅ All problematic images saved to: {all_txt_path}")
    
    # 各类型统计
    print(f"\nBreakdown by match type:")
    print("-" * 40)
    for match_type, problematic_set in problematic_by_type.items():
        if problematic_set:
            type_txt_path = os.path.join(output_dir, f"problematic_images_{match_type}_{timestamp}.txt")
            print(f"  {match_type.upper():<8}: {len(problematic_set):>4} images ({len(problematic_set)/len(gold_by_name)*100:>5.2f}%) -> {type_txt_path}")
    
    # 交叉分析
    print(f"\nCross-analysis:")
    print("-" * 40)
    only_hard = problematic_by_type["hard"] - problematic_by_type["soft"] - problematic_by_type["hybrid"]
    only_soft = problematic_by_type["soft"] - problematic_by_type["hard"] - problematic_by_type["hybrid"]
    only_hybrid = problematic_by_type["hybrid"] - problematic_by_type["hard"] - problematic_by_type["soft"]
    hard_soft = problematic_by_type["hard"] & problematic_by_type["soft"]
    hard_hybrid = problematic_by_type["hard"] & problematic_by_type["hybrid"]
    soft_hybrid = problematic_by_type["soft"] & problematic_by_type["hybrid"]
    all_three = problematic_by_type["hard"] & problematic_by_type["soft"] & problematic_by_type["hybrid"]
    
    print(f"  Only in HARD:     {len(only_hard):>4} images")
    print(f"  Only in SOFT:     {len(only_soft):>4} images")
    print(f"  Only in HYBRID:   {len(only_hybrid):>4} images")
    print(f"  HARD + SOFT:      {len(hard_soft):>4} images")
    print(f"  HARD + HYBRID:    {len(hard_hybrid):>4} images")
    print(f"  SOFT + HYBRID:    {len(soft_hybrid):>4} images")
    print(f"  ALL THREE:        {len(all_three):>4} images")
    
    # 显示样例
    if all_problematic:
        print(f"\nFirst 10 problematic images (any type):")
        for i, img_name in enumerate(sorted(all_problematic)[:10], 1):
            # 显示该图片在哪些类型中出错
            types = [t for t, s in problematic_by_type.items() if img_name in s]
            print(f"  {i:2d}. {img_name} (in: {', '.join(types)})")
        if len(all_problematic) > 10:
            print(f"  ... and {len(all_problematic) - 10} more")
    
    print("=" * 100 + "\n")


# =========================
# CLI
# =========================
def main():
    parser = argparse.ArgumentParser(description="Comprehensive evaluation and report generation for reaction extraction.")
    parser.add_argument("--ground_truth_file", type=str, required=True, help="Path to the ground truth JSON file.")
    parser.add_argument("--pred_file", type=str, required=True, help="Path to the prediction JSON file.")
    parser.add_argument("--image_base_path", type=str, required=True, help="Path to the directory containing images.")
    parser.add_argument("--output_dir", type=str, default="./eval_results", help="Directory to save reports and visualizations.")
    parser.add_argument("--filter_recover", action="store_true", help="Enable filter recovery mode.")
    parser.add_argument("--vis_range", type=int, nargs=2, metavar=("MIN", "MAX"), default=None, help="Visualize cases where MIN <= (FP + FN) <= MAX.")
    parser.add_argument("--limit_vis", type=int, default=-1, help="Limit number of visualizations in Markdown (<=0 means no limit).")
    parser.add_argument("--mode", type=str, default="all", choices=["all", "overall", "visualize", "report", "export_excel", "collect_problems", "export_f1_rank"], help="Run parts of the pipeline.")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    log_file = setup_logging(args.output_dir)
    logging.info(f"Start evaluation, log file: {log_file}")

    with open(args.ground_truth_file, "r", encoding="utf-8") as f:
        gold_data = json.load(f)
    with open(args.pred_file, "r", encoding="utf-8") as f:
        pred_data = json.load(f)

    gold_images = gold_data.get("images", [])
    pred_images = pred_data.get("images", [])
    if not gold_images:
        print("Error: No images found in ground truth file")
        return
    if not pred_images:
        print("Error: No images found in prediction file")
        return

    gold_names = {img.get("file_name", "") for img in gold_images}
    pred_names = {img.get("file_name", "") for img in pred_images}
    common = gold_names & pred_names
    only_gold = gold_names - pred_names
    only_pred = pred_names - gold_names

    print("Data summary:")
    print(f"  Ground truth images: {len(gold_images)}")
    print(f"  Prediction images:   {len(pred_images)}")
    print(f"  Common filenames:    {len(common)}")
    print(f"  Only in ground truth:{len(only_gold)}")
    print(f"  Only in predictions: {len(only_pred)}")

    if not common:
        print("Error: No matching filenames between ground truth and predictions")
        debug_filename_matching(gold_images, pred_images)
        return
    if only_gold:
        print(f"Warning: {len(only_gold)} ground truth images have no corresponding predictions")
    if only_pred:
        print(f"Warning: {len(only_pred)} prediction images have no corresponding ground truth")
    if len(common) < min(len(gold_names), len(pred_names)) * 0.5:
        print("Warning: <50% files matched by name. Running debug...")
        debug_filename_matching(gold_images, pred_images)

    validate_data_integrity(gold_images, pred_images)

    if args.mode in ["all", "overall", "report"]:
        run_overall_evaluation(gold_images, pred_images, args.image_base_path, filter_recover=args.filter_recover)

    if args.mode in ["all", "visualize"]:
        # Hard
        # generate_md_for_setting(
        #     gold_images, pred_images, args.image_base_path, args.output_dir, args.vis_range, args.limit_vis, mol_only=False, match_type="hard", filter_recover=args.filter_recover
        # )
        # Soft
        generate_md_for_setting(
            gold_images, pred_images, args.image_base_path, args.output_dir, args.vis_range, args.limit_vis, mol_only=True, match_type="soft", filter_recover=args.filter_recover
        )
        # Hybrid
        generate_md_for_setting(
            gold_images, pred_images, args.image_base_path, args.output_dir, args.vis_range, args.limit_vis, mol_only=False, match_type="hybrid", filter_recover=args.filter_recover
        )

    if args.mode in ["all", "report"]:
        run_itemized_evaluation(
            gold_images,
            pred_images,
            args.output_dir,
            args.image_base_path,
            vis_range=args.vis_range,
            limit_vis=args.limit_vis,
            generate_report=True,
            report_only_visualized=False,
            filter_recover=args.filter_recover,
        )

    if args.mode in ["all", "export_excel"]:
        save_match_results_to_excel(gold_images, pred_images, args.image_base_path, args.output_dir, args.filter_recover)

    if args.mode in ["all", "collect_problems"]:
        collect_problematic_images(gold_images, pred_images, args.image_base_path, args.output_dir, args.filter_recover)

    if args.mode in ["all", "export_f1_rank"]:
        save_per_image_f1_ranking_to_csv(gold_images, pred_images, args.image_base_path, args.output_dir, args.filter_recover)


if __name__ == "__main__":
    main()